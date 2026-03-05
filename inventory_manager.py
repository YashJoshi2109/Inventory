"""
Lab Inventory Manager — Python Backend
=======================================
Reads/writes Lab_Inventory_Barcode_System.xlsx using the exact sheet/column
structure from the Inventory_Barcode_System_Template.

Sheets:
  Items_Master      — SKU registry (cols A–L)
  Transactions      — Every IN/OUT log (cols A–H); G=In Qty, H=Out Qty
  Inventory_Summary — Auto-calculated by SUMIFS formulas (read-only via Python)
  SKU_Finder        — Single-SKU lookup (cell B2 = scan target)
  Graph             — Chart data feed
  Labels_5161       — Avery label sheet
  Dashboard         — KPI summary

Usage:
    python inventory_manager.py query
    python inventory_manager.py query --sku LAB-001
    python inventory_manager.py query --category Reagents --low-stock
    python inventory_manager.py receive --sku LAB-001 --qty 500 --notes "Restocked"
    python inventory_manager.py use --sku LAB-002 --qty 2 --loc 1 --notes "EXP-2026-005"
    python inventory_manager.py add-item --sku LAB-016 --name "Trypan Blue" --category Reagents --cost 18.50 --reorder 3
    python inventory_manager.py reorder
    python inventory_manager.py rfid --tag RFID-0003
    python inventory_manager.py sync
"""

import argparse
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import datetime
import json
import sys
from pathlib import Path

DB_PATH = Path(__file__).resolve().parent / "Lab_Inventory_Barcode_System.xlsx"

# Column indices (1-based), matching template exactly
# Items_Master: SKU(1) Desc(2) Cat(3) UnitCost(4) SalesPrice(5) ReorderLvl(6)
#               LeadDays(7) DateNeeded(8) Loc1Bin(9) Loc2Bin(10) Loc3Bin(11) BarcodeText(12)
IM = dict(sku=1, desc=2, cat=3, cost=4, price=5, reorder=6,
          lead=7, date_needed=8, loc1=9, loc2=10, loc3=11, barcode=12)

# Transactions: Date(1) Type(2) SKU(3) Qty(4) Location(5) Notes(6) InQty(7) OutQty(8)
TX = dict(date=1, type=2, sku=3, qty=4, loc=5, notes=6, in_qty=7, out_qty=8)


def thin_border():
    s = Side(style="thin", color="BBBBBB")
    return Border(top=s, bottom=s, left=s, right=s)

def hfill(hex_c):
    return PatternFill("solid", start_color=hex_c, fgColor=hex_c)

def row_fill(r):
    return "D6E4F0" if r % 2 == 0 else "FFFFFF"

def today_str():
    return datetime.date.today().isoformat()

def set_cell(ws, row, col, value, fmt=None, bold=False, center=False, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font   = Font(name="Calibri", size=10, bold=bold, color="000000")
    c.border = thin_border()
    c.alignment = Alignment(vertical="center", horizontal="center" if center else "left")
    if fmt:  c.number_format = fmt
    if fill: c.fill = hfill(fill)
    return c

def find_item_row(ws_master, sku):
    for row in ws_master.iter_rows(min_row=2, max_col=1):
        if row[0].value and str(row[0].value).strip().upper() == sku.upper():
            return row[0].row
    return None

def next_txn_row(ws_txn):
    for row in ws_txn.iter_rows(min_row=2, max_col=1):
        if row[0].value is None:
            return row[0].row
    return ws_txn.max_row + 1

def get_item_field(ws_master, inv_row, col_key):
    return ws_master.cell(row=inv_row, column=IM[col_key]).value

def read_items():
    df = pd.read_excel(DB_PATH, sheet_name="Items_Master", header=0)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(subset=["SKU"])
    df = df.rename(columns={
        "SKU": "sku", "Description": "desc", "Category": "category",
        "Unit Cost": "unit_cost", "Sales Price": "sales_price",
        "Reorder Level": "reorder_level", "Lead Time (days)": "lead_days",
        "Date Needed": "date_needed", "Loc1 Bin": "loc1_bin",
        "Loc2 Bin": "loc2_bin", "Loc3 Bin": "loc3_bin",
        "Barcode Text (SKU)": "barcode"
    })
    return df

def read_transactions():
    df = pd.read_excel(DB_PATH, sheet_name="Transactions", header=0)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(subset=["SKU"])
    df = df.rename(columns={
        "Date": "date", "Type (IN/OUT)": "type", "SKU": "sku",
        "Qty": "qty", "Location (1-3)": "location", "Notes": "notes",
        "In Qty": "in_qty", "Out Qty": "out_qty"
    })
    return df

def compute_on_hand(items_df, txn_df):
    """Mirrors the SUMIFS logic from Inventory_Summary sheet."""
    for loc in [1, 2, 3]:
        loc_in  = txn_df[txn_df["location"] == loc].groupby("sku")["in_qty"].sum()
        loc_out = txn_df[txn_df["location"] == loc].groupby("sku")["out_qty"].sum()
        items_df[f"loc{loc}_on_hand"] = (
            items_df["sku"].map(loc_in).fillna(0) -
            items_df["sku"].map(loc_out).fillna(0)
        )
    items_df["on_hand"] = (
        items_df["loc1_on_hand"] +
        items_df["loc2_on_hand"] +
        items_df["loc3_on_hand"]
    )
    items_df["scan_count"] = items_df["sku"].map(
        txn_df.groupby("sku")["qty"].count()
    ).fillna(0).astype(int)
    items_df["status"] = items_df.apply(
        lambda r: "LOW" if r["on_hand"] <= r["reorder_level"] else "OK", axis=1
    )
    return items_df


def query(args):
    items = read_items()
    txns  = read_transactions()
    df    = compute_on_hand(items, txns)

    if args.sku:       df = df[df["sku"].str.upper() == args.sku.upper()]
    if args.category:  df = df[df["category"].str.lower() == args.category.lower()]
    if args.status:    df = df[df["status"].str.upper() == args.status.upper()]
    if args.location:  df = df[df["loc1_bin"].str.lower().str.contains(args.location.lower(), na=False)]
    if args.low_stock: df = df[df["on_hand"] <= df["reorder_level"]]

    if df.empty:
        print("No matching items found.")
        return

    cols = ["sku","desc","category","on_hand","reorder_level","unit_cost","loc1_bin","status","scan_count"]
    print(df[cols].to_string(index=False))
    print(f"\n{len(df)} item(s) found.")
    if args.json:
        print(json.dumps(df[cols].fillna("").to_dict(orient="records"), default=str, indent=2))


def receive(args):
    wb       = openpyxl.load_workbook(DB_PATH)
    ws_items = wb["Items_Master"]
    ws_txn   = wb["Transactions"]

    inv_row = find_item_row(ws_items, args.sku)
    if inv_row is None:
        print(f"SKU '{args.sku}' not found. Use add-item first.")
        wb.close(); return

    r    = next_txn_row(ws_txn)
    fill = row_fill(r)
    qty  = float(args.qty)
    loc  = int(args.loc) if args.loc else 1

    set_cell(ws_txn, r, TX["date"],    datetime.datetime.now(), fmt="MM/DD/YYYY", center=True, fill=fill)
    set_cell(ws_txn, r, TX["type"],    "IN",  center=True, fill=fill)
    set_cell(ws_txn, r, TX["sku"],     args.sku.upper(), center=True, fill=fill)
    set_cell(ws_txn, r, TX["qty"],     qty,   fmt="0", center=True, fill=fill)
    set_cell(ws_txn, r, TX["loc"],     loc,   fmt="0", center=True, fill=fill)
    set_cell(ws_txn, r, TX["notes"],   args.notes or "Received", fill=fill)
    set_cell(ws_txn, r, TX["in_qty"],  qty,   fmt="0", center=True, fill=fill)
    set_cell(ws_txn, r, TX["out_qty"], 0,     fmt="0", center=True, fill=fill)
    ws_txn.row_dimensions[r].height = 18

    wb.save(DB_PATH)
    desc = get_item_field(ws_items, inv_row, "desc")
    print(f"IN  | {args.sku} | {desc} | +{qty} @ Loc{loc}")
    wb.close()


def use(args):
    wb       = openpyxl.load_workbook(DB_PATH)
    ws_items = wb["Items_Master"]
    ws_txn   = wb["Transactions"]

    inv_row = find_item_row(ws_items, args.sku)
    if inv_row is None:
        print(f"SKU '{args.sku}' not found in Items_Master.")
        wb.close(); return

    r    = next_txn_row(ws_txn)
    fill = row_fill(r)
    qty  = float(args.qty)
    loc  = int(args.loc) if args.loc else 1

    set_cell(ws_txn, r, TX["date"],    datetime.datetime.now(), fmt="MM/DD/YYYY", center=True, fill=fill)
    set_cell(ws_txn, r, TX["type"],    "OUT", center=True, fill=fill)
    set_cell(ws_txn, r, TX["sku"],     args.sku.upper(), center=True, fill=fill)
    set_cell(ws_txn, r, TX["qty"],     qty,   fmt="0", center=True, fill=fill)
    set_cell(ws_txn, r, TX["loc"],     loc,   fmt="0", center=True, fill=fill)
    set_cell(ws_txn, r, TX["notes"],   args.notes or "Used", fill=fill)
    set_cell(ws_txn, r, TX["in_qty"],  0,     fmt="0", center=True, fill=fill)
    set_cell(ws_txn, r, TX["out_qty"], qty,   fmt="0", center=True, fill=fill)
    ws_txn.row_dimensions[r].height = 18

    wb.save(DB_PATH)
    desc    = get_item_field(ws_items, inv_row, "desc")
    reorder = get_item_field(ws_items, inv_row, "reorder")
    print(f"OUT | {args.sku} | {desc} | -{qty} @ Loc{loc}")
    wb.close()

    # Reorder alert
    txns  = read_transactions()
    items = read_items()
    df    = compute_on_hand(items, txns)
    row   = df[df["sku"].str.upper() == args.sku.upper()]
    if not row.empty and row.iloc[0]["on_hand"] <= float(reorder or 0):
        print(f"ALERT: {args.sku} on-hand ({row.iloc[0]['on_hand']:.0f}) <= reorder level ({reorder}). Order needed!")


def add_item(args):
    wb       = openpyxl.load_workbook(DB_PATH)
    ws_items = wb["Items_Master"]

    if find_item_row(ws_items, args.sku):
        print(f"SKU '{args.sku}' already exists.")
        wb.close(); return

    r = 2
    while ws_items.cell(row=r, column=1).value:
        r += 1

    fill      = row_fill(r)
    lead      = int(args.lead or 7)
    date_need = datetime.datetime.now() + datetime.timedelta(days=lead)

    field_map = [
        (IM["sku"],         args.sku.upper(),            None),
        (IM["desc"],        args.name,                   None),
        (IM["cat"],         args.category or "Other",    None),
        (IM["cost"],        float(args.cost or 0),       "$#,##0.00"),
        (IM["price"],       float(args.price or 0),      "$#,##0.00"),
        (IM["reorder"],     float(args.reorder or 10),   "0"),
        (IM["lead"],        lead,                        "0"),
        (IM["date_needed"], date_need,                   "MM/DD/YYYY"),
        (IM["loc1"],        getattr(args, "loc1_bin", "") or "", None),
        (IM["loc2"],        getattr(args, "loc2_bin", "") or "", None),
        (IM["loc3"],        getattr(args, "loc3_bin", "") or "", None),
        (IM["barcode"],     args.sku.upper(),            None),
    ]
    for col, val, fmt in field_map:
        c = ws_items.cell(row=r, column=col, value=val)
        c.font      = Font(name="Calibri", size=10, color="000000")
        c.border    = thin_border()
        c.fill      = hfill(fill)
        c.alignment = Alignment(
            vertical="center",
            horizontal="center" if col in (4,5,6,7,8) else "left"
        )
        if fmt: c.number_format = fmt

    ws_items.row_dimensions[r].height = 18
    wb.save(DB_PATH)
    print(f"Added: {args.sku} — {args.name}")
    wb.close()


def reorder_report(args):
    items = read_items()
    txns  = read_transactions()
    df    = compute_on_hand(items, txns)
    low   = df[df["on_hand"] <= df["reorder_level"]].copy()
    low["deficit"] = low["reorder_level"] - low["on_hand"]

    if low.empty:
        print("All items above reorder level. No action needed.")
        return

    print(f"\n{'='*72}\n  REORDER REPORT  —  {today_str()}\n{'='*72}")
    cols = ["sku","desc","category","on_hand","reorder_level","deficit","unit_cost","loc1_bin","lead_days"]
    print(low[cols].to_string(index=False))
    print(f"\n{len(low)} item(s) need reordering.")

    if args.export:
        fname = f"reorder_report_{today_str()}.csv"
        low[cols].to_csv(fname, index=False)
        print(f"Exported to {fname}")


def rfid_lookup_cmd(args):
    sku   = args.tag.replace("RFID-", "LAB-").upper()
    items = read_items()
    txns  = read_transactions()
    df    = compute_on_hand(items, txns)
    row   = df[df["sku"].str.upper() == sku]

    if row.empty:
        print(f"No item found for tag '{args.tag}' (mapped SKU: {sku})")
        return

    item = row.iloc[0]
    print(f"\nRFID TAG:  {args.tag}")
    print(f"SKU:       {item['sku']}")
    print(f"Item:      {item['desc']}")
    print(f"Category:  {item['category']}")
    print(f"On Hand:   {item['on_hand']:.0f}")
    print(f"Status:    {item['status']}")
    print(f"Location:  {item['loc1_bin']}")
    print(f"Scans:     {item['scan_count']}")


def sync(args):
    items = read_items()
    txns  = read_transactions()
    df    = compute_on_hand(items, txns)

    print(f"\n{'='*60}")
    print(f"  DATABASE SYNC  —  {today_str()}")
    print(f"  File: {DB_PATH.name}")
    print(f"{'='*60}")
    print(f"  Total SKUs:          {len(df)}")
    print(f"  Total Transactions:  {len(txns)}")
    print(f"  Items OK:            {len(df[df['status']=='OK'])}")
    print(f"  Items LOW/Reorder:   {len(df[df['status']=='LOW'])}")
    print(f"  Total On-Hand Value: ${(df['on_hand'] * df['unit_cost']).sum():,.2f}")
    print(f"{'='*60}")

    low = df[df["status"] == "LOW"]
    if not low.empty:
        print("\nItems needing reorder:")
        for _, r in low.iterrows():
            print(f"   {r['sku']:12} {str(r['desc'])[:28]:28} on_hand={r['on_hand']:.0f}  reorder<={r['reorder_level']:.0f}")


def main():
    parser = argparse.ArgumentParser(
        description="Lab Inventory Manager — reads/writes Lab_Inventory_Barcode_System.xlsx"
    )
    sub = parser.add_subparsers(dest="command")

    pq = sub.add_parser("query")
    pq.add_argument("--sku"); pq.add_argument("--category"); pq.add_argument("--status")
    pq.add_argument("--location"); pq.add_argument("--low-stock", action="store_true", dest="low_stock")
    pq.add_argument("--json", action="store_true")

    pr = sub.add_parser("receive")
    pr.add_argument("--sku", required=True); pr.add_argument("--qty", required=True, type=float)
    pr.add_argument("--loc", default=1, type=int, choices=[1,2,3]); pr.add_argument("--notes", default="")

    pu = sub.add_parser("use")
    pu.add_argument("--sku", required=True); pu.add_argument("--qty", required=True, type=float)
    pu.add_argument("--loc", default=1, type=int, choices=[1,2,3]); pu.add_argument("--notes", default="")

    pa = sub.add_parser("add-item")
    pa.add_argument("--sku", required=True); pa.add_argument("--name", required=True)
    pa.add_argument("--category", default="Other"); pa.add_argument("--cost", type=float, default=0.0)
    pa.add_argument("--price", type=float, default=0.0); pa.add_argument("--reorder", type=float, default=10)
    pa.add_argument("--lead", type=int, default=7)
    pa.add_argument("--loc1-bin", dest="loc1_bin", default="")
    pa.add_argument("--loc2-bin", dest="loc2_bin", default="")
    pa.add_argument("--loc3-bin", dest="loc3_bin", default="")

    pro = sub.add_parser("reorder")
    pro.add_argument("--export", action="store_true")

    prf = sub.add_parser("rfid")
    prf.add_argument("--tag", required=True)

    sub.add_parser("sync")

    args = parser.parse_args()
    dispatch = {
        "query":    query,    "receive": receive, "use":     use,
        "add-item": add_item, "reorder": reorder_report,
        "rfid":     rfid_lookup_cmd, "sync": sync,
    }
    fn = dispatch.get(args.command)
    if fn: fn(args)
    else:  parser.print_help()

if __name__ == "__main__":
    main()
